import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


def analyze_data(df):
    total = len(df)
    completed = len(df[df['진행상태'] == '1차 완료']) if '진행상태' in df.columns else 0
    rate = completed / total if total > 0 else 0

    avg_sec = None
    if '처리소요초' in df.columns:
        valid = df['처리소요초'].dropna()
        avg_sec = valid.mean() if len(valid) > 0 else None

    error_dist = {}
    if '시스템/장비명' in df.columns:
        error_dist = df['시스템/장비명'].value_counts().to_dict()

    action_dist = {}
    if '조치유형' in df.columns:
        action_dist = df['조치유형'].value_counts().to_dict()

    company_dist = {}
    if '회사명' in df.columns:
        company_dist = df['회사명'].value_counts().to_dict()

    monthly_trend = {}
    if '통화시작시간' in df.columns:
        df2 = df.copy()
        df2['통화시작시간'] = pd.to_datetime(df2['통화시작시간'], errors='coerce')
        df2['월'] = df2['통화시작시간'].dt.to_period('M').astype(str)
        monthly_trend = df2['월'].value_counts().sort_index().to_dict()

    top_errors = list(error_dist.keys())[:5]

    return {
        'total_tickets': total,
        'completed_count': completed,
        'completion_rate': rate,
        'avg_resolution_seconds': avg_sec,
        'avg_resolution_str': _sec_to_str(avg_sec),
        'error_distribution': error_dist,
        'action_distribution': action_dist,
        'company_distribution': company_dist,
        'monthly_trend': monthly_trend,
        'top_errors': top_errors,
    }


def _sec_to_str(seconds):
    if seconds is None:
        return "-"
    try:
        s = int(seconds)
        h = s // 3600
        m = (s % 3600) // 60
        return f"{h}시간 {m}분"
    except:
        return "-"


def _header_style(ws, row=1):
    fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    for cell in ws[row]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)


def create_summary_sheet(wb, analysis):
    ws = wb.active
    ws.title = '요약'
    ws['A1'] = 'IT Helpdesk 월간 보고서 - 처리 현황 요약'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:C1')

    headers = ['항목', '값', '비고']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _header_style(ws, row=3)

    rows = [
        ('총 접수 건수', f"{analysis['total_tickets']:,}건", ''),
        ('1차 완료 건수', f"{analysis['completed_count']:,}건", ''),
        ('1차 처리율', f"{analysis['completion_rate']:.1%}", ''),
        ('평균 처리 시간', analysis['avg_resolution_str'], ''),
        ('주요 오류 시스템 Top3', ', '.join(analysis['top_errors'][:3]), ''),
    ]
    for r, (a, b, c) in enumerate(rows, 4):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)

    _auto_width(ws)
    return ws


def create_chart_sheet(wb, analysis):
    ws = wb.create_sheet('시스템별 현황')
    ws['A1'] = '시스템/장비별 접수 현황'
    ws['A1'].font = Font(bold=True, size=12, name='Arial')

    ws['A3'] = '시스템/장비명'
    ws['B3'] = '접수건수'
    ws['C3'] = '비율(%)'
    _header_style(ws, row=3)

    total = sum(analysis['error_distribution'].values()) or 1
    for r, (name, cnt) in enumerate(analysis['error_distribution'].items(), 4):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=f'{cnt/total:.1%}')

    _auto_width(ws)
    return ws


def create_detail_sheet(wb, df):
    ws = wb.create_sheet('상세 데이터')
    display_cols = [c for c in ['접수번호', '통화시작시간', '회사명', '요청자', '구분', '요청부서',
                                '요청유형', '시스템/장비명', '조치유형', '처리자', '조치내용',
                                '진행상태', '완료일시', '처리소요시간', '비고'] if c in df.columns]
    df_disp = df[display_cols].copy()

    # 헤더
    for c, col_name in enumerate(display_cols, 1):
        ws.cell(row=1, column=c, value=col_name)
    _header_style(ws, row=1)

    # 데이터
    for r, row in enumerate(df_disp.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=str(val) if pd.notna(val) else '')

    _auto_width(ws)
    return ws


def create_monthly_sheet(wb, analysis):
    ws = wb.create_sheet('월별 추이')
    ws['A1'] = '월별 접수 추이'
    ws['A1'].font = Font(bold=True, size=12, name='Arial')
    ws['A3'] = '월'
    ws['B3'] = '접수건수'
    _header_style(ws, row=3)

    for r, (month, cnt) in enumerate(sorted(analysis['monthly_trend'].items()), 4):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=cnt)

    _auto_width(ws)
    return ws


def generate_report(df, output_path=None):
    analysis = analyze_data(df)
    wb = openpyxl.Workbook()

    create_summary_sheet(wb, analysis)
    create_chart_sheet(wb, analysis)
    create_monthly_sheet(wb, analysis)
    create_detail_sheet(wb, df)

    if output_path:
        wb.save(output_path)
        return output_path
    else:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
