import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

st.set_page_config(
    page_title="IT Helpdesk 관리 시스템",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ═══════════════════════════════════════════════════════════════
# 데이터 처리 함수
# ═══════════════════════════════════════════════════════════════

def read_excel(file):
    try:
        df = pd.read_excel(file, sheet_name='관리대장', dtype=str)
    except:
        df = pd.read_excel(file, sheet_name=0, dtype=str)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({'nan': '', 'None': ''})
    return df

def parse_duration(s):
    try:
        parts = str(s).strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    except:
        pass
    return None

def sec_to_str(sec):
    if sec is None:
        return "-"
    try:
        h = int(sec) // 3600
        m = (int(sec) % 3600) // 60
        return f"{h}시간 {m}분"
    except:
        return "-"

def analyze(df):
    total = len(df)
    completed = len(df[df['진행상태'] == '1차 완료']) if '진행상태' in df.columns else 0
    rate = completed / total if total > 0 else 0
    avg_sec = None
    if '처리소요시간' in df.columns:
        secs = df['처리소요시간'].apply(parse_duration).dropna()
        if len(secs) > 0:
            avg_sec = secs.mean()
    error_dist = df['시스템/장비명'].value_counts().to_dict() if '시스템/장비명' in df.columns else {}
    action_dist = df['조치유형'].value_counts().to_dict() if '조치유형' in df.columns else {}
    company_dist = df['회사명'].value_counts().to_dict() if '회사명' in df.columns else {}
    monthly = {}
    if '통화시작시간' in df.columns:
        df2 = df.copy()
        df2['통화시작시간'] = pd.to_datetime(df2['통화시작시간'], errors='coerce')
        df2['월'] = df2['통화시작시간'].dt.to_period('M').astype(str)
        monthly = df2['월'].value_counts().sort_index().to_dict()
    return {
        'total': total, 'completed': completed, 'rate': rate,
        'avg_str': sec_to_str(avg_sec),
        'error_dist': error_dist, 'action_dist': action_dist,
        'company_dist': company_dist, 'monthly': monthly,
    }

def search(keyword, df, match='partial'):
    if not keyword or df is None or df.empty:
        return pd.DataFrame()
    cols = [c for c in ['요청유형', '시스템/장비명', '조치내용', '조치유형'] if c in df.columns]
    if not cols:
        return pd.DataFrame()
    if match == 'exact':
        mask = df[cols].apply(lambda c: c.astype(str).str.lower() == keyword.lower()).any(axis=1)
    else:
        mask = df[cols].apply(lambda c: c.astype(str).str.contains(keyword, case=False, na=False)).any(axis=1)
    return df[mask].copy()

def make_excel_report(df, analysis):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '요약'
    ws['A1'] = 'IT Helpdesk 월간 보고서'
    ws['A1'].font = Font(bold=True, size=14)
    for i, h in enumerate(['항목', '값'], 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    rows = [
        ('총 접수 건수', f"{analysis['total']:,}건"),
        ('1차 완료', f"{analysis['completed']:,}건"),
        ('처리율', f"{analysis['rate']:.1%}"),
        ('평균 처리시간', analysis['avg_str']),
    ]
    for i, (a, b) in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet('시스템별')
    ws2['A1'] = '시스템/장비별 접수 현황'
    ws2['A1'].font = Font(bold=True, size=12)
    for i, h in enumerate(['시스템/장비명', '건수'], 1):
        c = ws2.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for i, (k, v) in enumerate(sorted(analysis['error_dist'].items(), key=lambda x: -x[1]), 4):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 10

    ws3 = wb.create_sheet('상세데이터')
    show_cols = [c for c in ['접수번호', '통화시작시간', '회사명', '요청유형',
                              '시스템/장비명', '조치유형', '조치내용',
                              '처리자', '진행상태', '처리소요시간'] if c in df.columns]
    for i, h in enumerate(show_cols, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for ri, row in enumerate(df[show_cols].itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws3.cell(row=ri, column=ci, value=str(val) if val and val != '' else '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# 원본 양식 보고서 생성 함수
# ═══════════════════════════════════════════════════════════════

SYSTEMS = ['NAC','VMFORT','VPN','MPS','DLP','통합 VOC','통합HR','팀즈',
           '통합회계','NERP','NTMS','현장관리앱','다이소파트너','포탈',
           '장비','문의','기타','프로그램관리','POS']
ACTIONS = ['PC관리','기타','사용자환경장애','계정','사용문의','업무문의','구매문의']
TYPES   = ['임직원','매장','협력업체']
COMPANY_SHEET_MAP = {
    '아성다이소': '아성다이소',
    'HMP':        '아성hmp',
    '아성':       '아성',
    '아성솔루션': '아성솔루션',
    '한웰':       '한웰',
    'HS':         'hs',
}

def _parse_sec(val):
    if not val or str(val).strip() in ['', 'nan', 'None']:
        return None
    val = str(val).strip()
    try:
        if ':' in val:
            p = val.split(':')
            if len(p) == 3:
                return int(p[0])*3600 + int(p[1])*60 + float(p[2])
        return float(val) * 24 * 3600
    except:
        return None

def _calc_stats(df, total_company=None):
    total = len(df)
    comp1 = len(df[df['진행상태'] == '1차 완료']) if '진행상태' in df.columns else 0
    comp2 = len(df[df['진행상태'] == '2차 완료']) if '진행상태' in df.columns else 0
    rate = comp1 / total if total > 0 else 0

    dates = pd.to_datetime(df['통화시작시간'], errors='coerce').dropna() if '통화시작시간' in df.columns else pd.Series([])
    if len(dates) > 0:
        period = f"접수기간 : {dates.min().strftime('%Y.%m.%d')} ~ {dates.max().strftime('%Y.%m.%d')}"
        d = dates.max()
        base_month = f"{d.year}년 {d.month:02d}월"
    else:
        period = "접수기간 : - ~ -"
        base_month = ""
    write_date = datetime.today().strftime('%Y.%m.%d')

    ref_total = total_company if total_company is not None else total
    sys_stats = []
    for sys in SYSTEMS:
        sdf = df[df['시스템/장비명'] == sys] if '시스템/장비명' in df.columns else pd.DataFrame()
        cnt = len(sdf)
        c1 = len(sdf[sdf['진행상태'] == '1차 완료']) if '진행상태' in sdf.columns and len(sdf) > 0 else 0
        r = (c1/cnt if cnt > 0 else 0) if total_company is None else (c1/ref_total if ref_total > 0 else 0)
        sys_stats.append((sys, cnt, c1, r))

    act_stats = []
    for act in ACTIONS:
        adf = df[df['조치유형'] == act] if '조치유형' in df.columns else pd.DataFrame()
        cnt = len(adf)
        c1 = len(adf[adf['진행상태'] == '1차 완료']) if '진행상태' in adf.columns and len(adf) > 0 else 0
        act_stats.append((act, cnt, c1, c1/cnt if cnt > 0 else 0))

    type_stats = []
    for t in TYPES:
        tdf = df[df['구분'] == t] if '구분' in df.columns else pd.DataFrame()
        cnt = len(tdf)
        c1 = len(tdf[tdf['진행상태'] == '1차 완료']) if '진행상태' in tdf.columns and len(tdf) > 0 else 0
        type_stats.append((t, cnt, c1, c1/cnt if cnt > 0 else 0))

    df2 = df.copy()
    df2['_초'] = df2['처리소요시간'].apply(_parse_sec) if '처리소요시간' in df2.columns else 0
    valid = df2[df2['_초'].notna()]
    grade_stats = [
        ('A (10분이내)', len(valid[valid['_초'] <= 600])),
        ('B (10~20분)',  len(valid[(valid['_초'] > 600) & (valid['_초'] <= 1200)])),
        ('C (20분초과)', len(valid[valid['_초'] > 1200])),
    ]
    grade_total = sum(g[1] for g in grade_stats)

    return {
        'total': total, 'comp1': comp1, 'comp2': comp2, 'rate': rate,
        'period': period, 'base_month': base_month, 'write_date': write_date,
        'sys_stats': sys_stats, 'act_stats': act_stats,
        'type_stats': type_stats, 'grade_stats': grade_stats,
        'grade_total': grade_total,
        'grade_total_r': grade_total/total if total > 0 else 0,
    }

def _fill_sheet(ws, s):
    ws['L1'] = s['base_month']
    ws['L3'] = s['write_date']
    ws['L4'] = f"{s['total']}건"
    ws['A8'] = f"{s['total']}건"
    ws['F8'] = f"{s['comp1']}건"
    ws['K8'] = s['rate']
    ws['A9'] = s['period']
    ws['F9'] = s['rate']
    ws['K9'] = f"{s['comp2']}건"
    ws['F74'] = s['total']
    ws['J74'] = s['comp1']
    ws['M74'] = s['comp2']
    ws['N74'] = s['rate']

    for i, (_, cnt, c1, r) in enumerate(s['sys_stats']):
        row = 77 + i
        ws[f'D{row}'] = cnt
        ws[f'H{row}'] = c1
        ws[f'L{row}'] = r
    sys_total = sum(x[1] for x in s['sys_stats'])
    sys_c1    = sum(x[2] for x in s['sys_stats'])
    ws['D96'] = sys_total
    ws['H96'] = sys_c1
    ws['L96'] = sys_c1/sys_total if sys_total > 0 else 0

    for i, (_, cnt, c1, r) in enumerate(s['act_stats']):
        ws[f'D{99+i}'] = cnt
        ws[f'H{99+i}'] = c1
        ws[f'L{99+i}'] = r
    act_total = sum(x[1] for x in s['act_stats'])
    act_c1    = sum(x[2] for x in s['act_stats'])
    ws['D106'] = act_total
    ws['H106'] = act_c1
    ws['L106'] = act_c1/act_total if act_total > 0 else 0

    for i, (_, cnt, c1, r) in enumerate(s['type_stats']):
        ws[f'D{109+i}'] = cnt
        ws[f'H{109+i}'] = c1
        ws[f'L{109+i}'] = r
    type_total = sum(x[1] for x in s['type_stats'])
    type_c1    = sum(x[2] for x in s['type_stats'])
    ws['D112'] = type_total
    ws['H112'] = type_c1
    ws['L112'] = type_c1/type_total if type_total > 0 else 0

    for i, (_, cnt) in enumerate(s['grade_stats']):
        r = cnt/s['total'] if s['total'] > 0 else 0
        ws[f'D{115+i}'] = cnt
        ws[f'H{115+i}'] = r
    ws['D118'] = s['grade_total']
    ws['H118'] = s['grade_total_r']

def generate_template_report(df_ledger, template_file):
    wb = load_workbook(template_file)
    stats_all = _calc_stats(df_ledger)
    if '총 접수현황' in wb.sheetnames:
        _fill_sheet(wb['총 접수현황'], stats_all)
    for company, sheet_name in COMPANY_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        cdf = df_ledger[df_ledger['회사명'] == company] if '회사명' in df_ledger.columns else pd.DataFrame()
        stats_c = _calc_stats(cdf, total_company=len(cdf))
        _fill_sheet(wb[sheet_name], stats_c)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════
# 사이드바
# ═══════════════════════════════════════════════════════════════

with st.sidebar:
    st.title("🔧 IT Helpdesk")
    st.markdown("---")
    st.subheader("📂 데이터 업로드")
    uploaded = st.file_uploader("관리대장 Excel 업로드", type=["xlsx", "xls"])
    if uploaded:
        try:
            df = read_excel(uploaded)
            st.session_state["df"] = df
            st.success(f"✅ 로드 완료! ({len(df):,}건)")
        except Exception as e:
            st.error(f"❌ 오류: {str(e)}")
    st.markdown("---")
    st.markdown("**📖 사용 가이드**")
    st.markdown("""
1. Excel 파일 업로드
2. **🔍 검색** 탭 → 오류 검색
3. **📊 보고서** 탭 → 통계 보고서
4. **📋 양식 보고서** 탭 → 원본 양식 그대로 출력
""")
    st.caption("IT Helpdesk 관리 시스템 v2.0")


# ═══════════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════════

st.title("🔧 IT Helpdesk 관리 시스템")
st.markdown("Excel 관리대장 자동 분석 · 오류 검색 · 보고서 생성")
st.markdown("---")

df = st.session_state.get("df")

if df is None:
    col1, col2, col3 = st.columns(3)
    col1.info("**🔍 오류 검색**\n\nNAC, VPN 등 키워드로 과거 사례 검색")
    col2.info("**📊 자동 보고서**\n\n월간 현황, 통계 자동 생성")
    col3.info("**📋 양식 보고서**\n\n기존 보고서 양식 그대로 수치 채워서 다운로드")
    st.markdown("---")
    st.markdown("### 👆 왼쪽에서 Excel 파일을 업로드하세요!")
    st.stop()

total = len(df)
completed = len(df[df['진행상태'] == '1차 완료']) if '진행상태' in df.columns else 0
rate = completed / total if total > 0 else 0

col1, col2, col3, col4 = st.columns(4)
col1.metric("📋 총 접수", f"{total:,}건")
col2.metric("✅ 1차 완료", f"{completed:,}건")
col3.metric("📈 처리율", f"{rate:.1%}")
col4.metric("🏢 고객사", f"{df['회사명'].nunique() if '회사명' in df.columns else '-'}개")
st.markdown("---")

tab1, tab2, tab3 = st.tabs(["🔍 검색", "📊 보고서", "📋 양식 보고서"])

# ── 검색 탭 ────────────────────────────────────────────────────
with tab1:
    st.subheader("🔍 오류 검색")
    col1, col2 = st.columns([3, 1])
    with col1:
        keyword = st.text_input("검색어 입력", placeholder="예: NAC, VPN, 팀즈, 계정...")
    with col2:
        match_type = st.radio("검색 방식", ["부분 일치", "정확 일치"], horizontal=True)

    st.markdown("**⚡ 빠른 검색:**")
    quick_terms = ['NAC', 'VPN', '팀즈', 'NERP', '장비', '계정', 'PC관리', '기타']
    q_cols = st.columns(len(quick_terms))
    for i, term in enumerate(quick_terms):
        if q_cols[i].button(term, use_container_width=True):
            keyword = term

    if keyword:
        match_code = 'partial' if match_type == "부분 일치" else 'exact'
        results = search(keyword, df, match_code)
        if results.empty:
            st.info(f"'{keyword}' 검색 결과가 없습니다.")
        else:
            st.success(f"✅ {len(results):,}건 검색됨")
            show = [c for c in ['접수번호', '통화시작시간', '회사명', '요청유형',
                                 '시스템/장비명', '조치유형', '조치내용',
                                 '처리자', '처리소요시간', '진행상태'] if c in results.columns]
            st.dataframe(results[show].reset_index(drop=True), use_container_width=True, height=400)
            csv = results[show].to_csv(index=False, encoding='utf-8-sig')
            st.download_button("⬇️ CSV 다운로드", data=csv.encode('utf-8-sig'),
                               file_name=f"검색_{keyword}.csv", mime="text/csv")
    else:
        if '시스템/장비명' in df.columns:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**시스템/장비별 Top10**")
                top10 = df['시스템/장비명'].value_counts().head(10).reset_index()
                top10.columns = ['시스템', '건수']
                st.dataframe(top10, use_container_width=True, hide_index=True)
            with col2:
                if '조치유형' in df.columns:
                    st.markdown("**조치유형별 현황**")
                    act = df['조치유형'].value_counts().reset_index()
                    act.columns = ['조치유형', '건수']
                    st.dataframe(act, use_container_width=True, hide_index=True)

# ── 보고서 탭 ──────────────────────────────────────────────────
with tab2:
    st.subheader("📊 자동 보고서 생성")
    companies = ['전체'] + sorted(df['회사명'].dropna().unique().tolist()) if '회사명' in df.columns else ['전체']
    selected = st.selectbox("고객사 선택", companies)
    df_f = df[df['회사명'] == selected].copy() if selected != '전체' else df.copy()
    st.markdown(f"분석 대상: **{len(df_f):,}건**")

    if st.button("📈 보고서 생성", type="primary"):
        with st.spinner("분석 중..."):
            a = analyze(df_f)
        st.success("✅ 완료!")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("총 접수", f"{a['total']:,}건")
        col2.metric("1차 완료", f"{a['completed']:,}건")
        col3.metric("처리율", f"{a['rate']:.1%}")
        col4.metric("평균 처리시간", a['avg_str'])
        st.markdown("---")

        if a['error_dist']:
            st.markdown("**시스템/장비별 현황**")
            edf = pd.DataFrame(list(a['error_dist'].items()), columns=['시스템', '건수']).sort_values('건수', ascending=False)
            col1, col2 = st.columns(2)
            col1.dataframe(edf.reset_index(drop=True), use_container_width=True, hide_index=True)
            col2.bar_chart(edf.set_index('시스템')['건수'])

        if a['monthly']:
            st.markdown("**월별 추이**")
            mdf = pd.DataFrame(list(a['monthly'].items()), columns=['월', '건수']).sort_values('월')
            st.line_chart(mdf.set_index('월')['건수'])

        if a['company_dist']:
            st.markdown("**고객사별 현황**")
            cdf = pd.DataFrame(list(a['company_dist'].items()), columns=['고객사', '건수']).sort_values('건수', ascending=False)
            col1, col2 = st.columns(2)
            col1.dataframe(cdf.reset_index(drop=True), use_container_width=True, hide_index=True)
            col2.bar_chart(cdf.set_index('고객사')['건수'])

        st.markdown("---")
        with st.spinner("Excel 생성 중..."):
            buf = make_excel_report(df_f, a)
        fname = f"IT_Helpdesk_보고서_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button("📥 Excel 보고서 다운로드", data=buf, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")

# ── 양식 보고서 탭 ─────────────────────────────────────────────
with tab3:
    st.subheader("📋 원본 양식 보고서 생성")
    st.markdown("""
기존에 사용하던 **월간 점검보고서 양식 파일**을 업로드하면,
관리대장 데이터의 수치를 자동으로 채워서 다운로드할 수 있어요.

> 양식 파일: `ITHELPDESK_월간보고서_완성.xlsx` 같은 형태의 파일
""")
    template_file = st.file_uploader(
        "📂 원본 보고서 양식 업로드",
        type=["xlsx"],
        key="template_uploader",
        help="기존 월간보고서 양식 파일을 업로드하세요. 서식/차트는 그대로 유지됩니다."
    )

    if template_file:
        st.success("✅ 양식 파일 로드됨!")
        if st.button("📋 양식 보고서 생성", type="primary"):
            with st.spinner("수치 계산 및 양식 채우는 중..."):
                try:
                    buf = generate_template_report(df, template_file)
                    fname = f"IT_Helpdesk_월간보고서_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    st.success("✅ 보고서 생성 완료!")
                    st.download_button(
                        "📥 월간보고서 다운로드",
                        data=buf,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )

                    # 미리보기
                    with st.expander("📊 생성된 수치 미리보기"):
                        s = _calc_stats(df)
                        col1, col2, col3 = st.columns(3)
                        col1.metric("총 접수", f"{s['total']:,}건")
                        col2.metric("1차 완료", f"{s['comp1']:,}건")
                        col3.metric("처리율", f"{s['rate']:.1%}")
                        st.caption(s['period'])

                except Exception as e:
                    st.error(f"❌ 오류 발생: {str(e)}")
                    st.info("양식 파일이 올바른 형식인지 확인해주세요.")
    else:
        st.info("👆 원본 보고서 양식 파일을 업로드하면 수치가 자동으로 채워진 보고서가 생성됩니다.")
        st.markdown("""
**지원하는 양식 구조:**
- `총 접수현황` 시트 (전체 현황)
- `아성다이소`, `아성hmp`, `아성` 등 고객사별 시트
- 처리 현황 요약, 시스템별/조치유형별/접수유형별/등급별 현황 자동 입력
        """)
